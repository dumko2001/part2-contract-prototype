#!/usr/bin/env python3

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

EPS = 1e-6


def read_rows(path: Path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        out.append(tuple(row[:10]))
    return out


def norm(v):
    if isinstance(v, (float, int)):
        return round(v, 2)
    return v


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser()
    parser.add_argument("--expected", default=str(base_dir / "FILE 3 OUTPUT.xlsx"))
    parser.add_argument("--actual", default=str(base_dir / "artifacts/FILE_3_GENERATED.xlsx"))
    args = parser.parse_args()

    exp = [tuple(norm(x) for x in r) for r in read_rows(Path(args.expected))]
    act = [tuple(norm(x) for x in r) for r in read_rows(Path(args.actual))]

    if exp == act:
        print("MATCH: Generated output matches expected File 3.")
        return 0

    print("MISMATCH: Differences found.")
    max_len = max(len(exp), len(act))
    for i in range(max_len):
        er = exp[i] if i < len(exp) else None
        ar = act[i] if i < len(act) else None
        if er != ar:
            print(f"Row {i+2}:")
            print(f"  expected={er}")
            print(f"  actual  ={ar}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
