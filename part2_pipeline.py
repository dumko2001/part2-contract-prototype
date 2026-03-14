#!/usr/bin/env python3
"""
Part 2 prototype:
Excel (File 1) + PDF (File 2) -> generated Excel output (File 3 format)

Architecture:
- Lane A: deterministic parser (default)
- Lane B: Groq structured output fallback for ambiguous instructions
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import textwrap
from dataclasses import dataclass, asdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from dotenv import load_dotenv
from groq import Groq
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


HEADER = [
    "Specification item number",
    "Description",
    "Unit",
    "Quantity",
    "Remove / Break up",
    "Dispose",
    "Reuse",
    "Clean",
    "Package",
    "Transport",
]

EPS = 1e-6


@dataclass
class Rule:
    pct: float
    basis: str = "quantity"  # quantity | reuse


@dataclass
class ParseResult:
    item_code: str
    instruction: str
    lane: str
    confidence: float
    remove: Rule
    dispose: Rule
    reuse: Rule
    clean: Rule
    package: Rule
    transport: Rule
    notes: List[str]
    lane_reason: str = ""


@dataclass
class ChildItem:
    code: str
    parent: str
    description: str
    unit: str
    quantity: float


@dataclass
class ParentItem:
    code: str
    description: str
    unit: str


def normalize_item_code(value) -> Optional[str]:
    if value is None:
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw = raw.replace(",", ".")
    if raw.endswith(".0"):
        raw = raw[:-2]
    if re.fullmatch(r"\d+(\.\d+)?", raw):
        return raw.split(".")[0]
    digits = re.sub(r"\D", "", raw)
    return digits if digits else None


def to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (float, int)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def load_file1(file1: Path) -> Tuple[Dict[str, ParentItem], Dict[str, ChildItem], List[str]]:
    wb = load_workbook(file1, data_only=True)
    ws = wb[wb.sheetnames[0]]

    parents: Dict[str, ParentItem] = {}
    children: Dict[str, ChildItem] = {}
    order: List[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = normalize_item_code(row[0])
        if not code:
            continue
        description = str(row[1] or "").strip()
        unit = str(row[2] or "").strip()
        qty = to_float(row[3])

        if len(code) == 4:
            parents[code] = ParentItem(code=code, description=description, unit=unit)
            continue
        if len(code) >= 6 and qty is not None:
            parent = code[:4]
            children[code] = ChildItem(
                code=code,
                parent=parent,
                description=description,
                unit=unit,
                quantity=float(qty),
            )
            order.append(code)

    return parents, children, order


def extract_pdf_text(pdf_path: Path) -> str:
    try:
        out = subprocess.check_output(
            ["pdftotext", "-layout", str(pdf_path), "-"],
            stderr=subprocess.STDOUT,
            text=True,
        )
        return out
    except (FileNotFoundError, subprocess.CalledProcessError) as exc:
        raise RuntimeError(
            "Could not extract PDF text. Ensure `pdftotext` is installed."
        ) from exc


def extract_instruction_blocks(pdf_text: str) -> Dict[str, dict]:
    # Keep page boundaries from pdftotext form-feed markers.
    lines = pdf_text.replace("\r\n", "\n").split("\n")
    blocks: Dict[str, List[Tuple[int, str]]] = {}
    current_code: Optional[str] = None
    page_no = 1

    for line in lines:
        if "\f" in line:
            parts = line.split("\f")
            # Process content before the last form-feed on current page.
            for i, part in enumerate(parts):
                if part:
                    m = re.match(r"^\s*(\d{6})\b", part)
                    if m:
                        current_code = m.group(1)
                        blocks[current_code] = [(page_no, part)]
                    elif current_code:
                        blocks[current_code].append((page_no, part))
                if i < len(parts) - 1:
                    page_no += 1
            continue
        m = re.match(r"^\s*(\d{6})\b", line)
        if m:
            current_code = m.group(1)
            blocks[current_code] = [(page_no, line)]
            continue
        if current_code:
            blocks[current_code].append((page_no, line))

    result: Dict[str, dict] = {}
    for code, rows in blocks.items():
        block_text = "\n".join(line for _, line in rows)
        first_page = rows[0][0] if rows else None
        ins_start = re.search(r"Instruction:\s*", block_text, flags=re.IGNORECASE)
        if not ins_start:
            continue
        # Reliable boundary: from "Instruction:" to end of this item's bounded block.
        instruction_raw = block_text[ins_start.end() :]
        instruction = re.sub(r"\s+", " ", instruction_raw).strip()
        block_hash = hashlib.sha256(block_text.encode("utf-8")).hexdigest()[:16]
        result[code] = {
            "instruction": instruction,
            "page": first_page,
            "block_hash": block_hash,
            "has_action_signal": bool(
                re.search(r"\b(reus|dispos|clean|packag|transport)\b|\d+(?:\.\d+)?\s*%", instruction, flags=re.IGNORECASE)
            ),
        }
    return result


def _find_pct(pattern: str, text: str) -> Optional[float]:
    m = re.search(pattern, text, flags=re.IGNORECASE)
    if not m:
        return None
    return float(m.group(1))


def deterministic_parse(item_code: str, instruction: str) -> ParseResult:
    text = instruction.lower()
    notes: List[str] = []

    reuse = Rule(0.0, "quantity")
    dispose = Rule(0.0, "quantity")
    clean = Rule(0.0, "quantity")
    package = Rule(0.0, "quantity")
    transport = Rule(0.0, "quantity")
    remove = Rule(100.0, "quantity")

    if re.search(r"all .*100%.*reus|all materials.*reused", text):
        reuse.pct = 100.0
        notes.append("reuse_all")
    rp = _find_pct(r"(\d+(?:\.\d+)?)\s*%[^.]*reus", text)
    if rp is not None:
        reuse.pct = max(reuse.pct, rp)
        notes.append("reuse_pct")

    if re.search(r"all .*100%.*dispos|must be disposed of", text):
        if "remaining" not in text:
            dp = _find_pct(r"(\d+(?:\.\d+)?)\s*%[^.]*dispos", text)
            dispose.pct = 100.0 if dp is None else max(100.0, dp)
            notes.append("dispose_all")
    dp = _find_pct(r"(\d+(?:\.\d+)?)\s*%[^.]*dispos", text)
    if dp is not None:
        dispose.pct = max(dispose.pct, dp)
        notes.append("dispose_pct")
    rem = _find_pct(r"remaining\s+(\d+(?:\.\d+)?)\s*%[^.]*dispos", text)
    if rem is not None:
        dispose.pct = rem
        notes.append("dispose_remaining")

    if re.search(r"fully\s*\(?100%\)?\s*clean|100%[^.]*clean", text):
        clean.pct = 100.0
        notes.append("clean_all")
    cp = _find_pct(r"(\d+(?:\.\d+)?)\s*%[^.]*clean", text)
    if cp is not None:
        clean.pct = max(clean.pct, cp)
        notes.append("clean_pct")
    if re.search(r"half of this|of this\s*\(50%\)", text):
        clean.pct = 50.0
        clean.basis = "reuse"
        notes.append("clean_half_of_reuse")

    pp = _find_pct(r"(\d+(?:\.\d+)?)\s*%[^.]*packag", text)
    if pp is not None:
        package.pct = pp
        notes.append("package_pct")
    elif "packag" in text and reuse.pct > 0:
        package.pct = reuse.pct
        notes.append("package_equals_reuse")

    if "transport cost" in text or "transport costs" in text:
        transport.pct = 100.0
        notes.append("transport_all")

    confidence = 1.0
    if reuse.pct == 0 and dispose.pct == 0 and clean.pct == 0 and package.pct == 0 and transport.pct == 0:
        confidence = 0.0
        notes.append("no_action_detected")
    if reuse.pct + dispose.pct > 100.0 + EPS:
        confidence = min(confidence, 0.3)
        notes.append("over_100_split")

    return ParseResult(
        item_code=item_code,
        instruction=instruction,
        lane="A",
        confidence=confidence,
        remove=remove,
        dispose=dispose,
        reuse=reuse,
        clean=clean,
        package=package,
        transport=transport,
        notes=notes,
        lane_reason="deterministic_parser_confident",
    )


def enforce_hard_rules(parsed: ParseResult) -> ParseResult:
    text = parsed.instruction.lower()

    # "transport costs charged" means transport applies to the full quantity.
    if "transport cost" in text or "transport costs" in text:
        parsed.transport = Rule(100.0, "quantity")
        parsed.notes.append("hard_rule_transport_all")

    # "remaining X% disposed" is explicit.
    rem = _find_pct(r"remaining\s+(\d+(?:\.\d+)?)\s*%[^.]*dispos", text)
    if rem is not None:
        parsed.dispose = Rule(rem, "quantity")
        parsed.notes.append("hard_rule_remaining_dispose")
        if parsed.reuse.pct <= EPS:
            parsed.reuse = Rule(max(0.0, 100.0 - rem), "quantity")
            parsed.notes.append("hard_rule_reuse_from_remaining")

    # Packaging is typically expressed as a share of the original quantity.
    if "packag" in text and parsed.package.pct > 0 and "of reuse" not in text and "of reused" not in text:
        parsed.package = Rule(parsed.package.pct, "quantity")
        parsed.notes.append("hard_rule_package_quantity_basis")

    if "packag" in text and parsed.package.pct <= EPS and parsed.reuse.pct > 0:
        parsed.package = Rule(parsed.reuse.pct, "quantity")
        parsed.notes.append("hard_rule_package_equals_reuse")

    return parsed


def groq_parse(
    item_code: str,
    instruction: str,
    model: str,
    api_key: str,
    timeout: int = 25,
) -> ParseResult:
    schema = {
        "name": "instruction_parse",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "reuse_pct": {"type": "number"},
                "dispose_pct": {"type": "number"},
                "clean_pct": {"type": "number"},
                "clean_basis": {"type": "string", "enum": ["quantity", "reuse"]},
                "package_pct": {"type": "number"},
                "package_basis": {"type": "string", "enum": ["quantity", "reuse"]},
                "transport_pct": {"type": "number"},
                "confidence": {"type": "number"},
                "notes": {"type": "array", "items": {"type": "string"}},
            },
            "required": [
                "reuse_pct",
                "dispose_pct",
                "clean_pct",
                "clean_basis",
                "package_pct",
                "package_basis",
                "transport_pct",
                "confidence",
                "notes",
            ],
            "additionalProperties": False,
        },
    }

    messages = [
        {
            "role": "system",
            "content": (
                "Extract action percentages from one construction instruction. "
                "Return only structured data. "
                "Percentages are 0..100 and represent shares of quantity unless basis says reuse."
            ),
        },
        {
            "role": "user",
            "content": f"Item {item_code}\nInstruction: {instruction}",
        },
    ]

    client = Groq(api_key=api_key)

    try:
        completion = client.chat.completions.create(
            model=model,
            temperature=0,
            messages=messages,
            response_format={"type": "json_schema", "json_schema": schema},
        )
        mode = "json_schema"
        msg = completion.choices[0].message.content
        if not msg:
            raise RuntimeError("Groq returned empty content in json_schema mode.")
        data = json.loads(msg)
    except Exception as exc:  # noqa: BLE001
        detail = str(exc)
        if ("json_schema" in detail) or ("error code: 1010" in detail):
            # For models that don't support json_schema (e.g. llama-3.1-8b-instant),
            # fall back to json_object and enforce schema locally.
            completion = client.chat.completions.create(
                model=model,
                temperature=0,
                messages=[
                    *messages,
                    {
                        "role": "user",
                        "content": (
                            "Return a JSON object with exactly these keys: "
                            "reuse_pct,dispose_pct,clean_pct,clean_basis,package_pct,"
                            "package_basis,transport_pct,confidence,notes"
                        ),
                    },
                ],
                response_format={"type": "json_object"},
            )
            mode = "json_object"
            msg = completion.choices[0].message.content
            if not msg:
                raise RuntimeError("Groq returned empty content in json_object mode.")
            data = json.loads(msg)
        else:
            raise RuntimeError(f"Groq request failed: {detail}") from exc

    def clamp(x: float) -> float:
        return max(0.0, min(100.0, float(x)))

    required = {
        "reuse_pct",
        "dispose_pct",
        "clean_pct",
        "clean_basis",
        "package_pct",
        "package_basis",
        "transport_pct",
        "confidence",
        "notes",
    }
    missing = sorted(required - set(data.keys()))
    if missing:
        raise RuntimeError(f"Groq fallback response missing keys: {missing}")

    clean_basis = data["clean_basis"] if data["clean_basis"] in {"quantity", "reuse"} else "quantity"
    package_basis = data["package_basis"] if data["package_basis"] in {"quantity", "reuse"} else "quantity"

    return ParseResult(
        item_code=item_code,
        instruction=instruction,
        lane=f"B:{mode}",
        confidence=float(data["confidence"]),
        remove=Rule(100.0, "quantity"),
        dispose=Rule(clamp(data["dispose_pct"]), "quantity"),
        reuse=Rule(clamp(data["reuse_pct"]), "quantity"),
        clean=Rule(clamp(data["clean_pct"]), clean_basis),
        package=Rule(clamp(data["package_pct"]), package_basis),
        transport=Rule(clamp(data["transport_pct"]), "quantity"),
        notes=list(data.get("notes", [])),
        lane_reason="fallback_llm_structured_output",
    )


def decide_lane_b_reason(parsed: ParseResult) -> Optional[str]:
    text = parsed.instruction.lower()
    if parsed.confidence < 0.8:
        return "low_confidence_deterministic_parse"
    if "remaining" in text and "dispose" in text and "dispose_remaining" not in parsed.notes:
        return "remaining_clause_not_resolved"
    if parsed.reuse.pct + parsed.dispose.pct > 100.0 + EPS:
        return "split_over_100"
    if parsed.reuse.pct == 0 and parsed.dispose.pct == 0 and parsed.clean.pct == 0 and parsed.package.pct == 0 and parsed.transport.pct == 0:
        return "no_action_detected"
    return None


def compute_amount(quantity: float, rule: Rule, reuse_amount: float) -> float:
    base = quantity
    if rule.basis == "reuse":
        base = reuse_amount
    return base * rule.pct / 100.0


def round2(x: float) -> float:
    return float(Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def blank_if_zero(x: float):
    rx = round2(x)
    if abs(rx) <= EPS:
        return None
    return rx


def process(
    file1: Path,
    file2_pdf: Path,
    output_xlsx: Path,
    report_path: Path,
    groq_model: str,
    groq_api_key: Optional[str],
    force_lane_b: bool = False,
    ambiguous_policy: str = "continue",
) -> None:
    parents, children, child_order = load_file1(file1)
    pdf_text = extract_pdf_text(file2_pdf)
    instructions = extract_instruction_blocks(pdf_text)

    per_child: Dict[str, ParseResult] = {}
    errors: List[dict] = []

    for code in child_order:
        child = children[code]
        instruction_meta = instructions.get(code)
        if not instruction_meta:
            errors.append({"item_code": code, "error": "instruction_not_found"})
            continue
        instruction = instruction_meta["instruction"]
        parsed = deterministic_parse(code, instruction)
        lane_b_reason = decide_lane_b_reason(parsed)
        use_lane_b = force_lane_b or (lane_b_reason is not None)
        if use_lane_b and groq_api_key:
            try:
                parsed = groq_parse(code, instruction, groq_model, groq_api_key)
                parsed.lane_reason = "forced_lane_b" if force_lane_b else lane_b_reason or "fallback_needed"
            except Exception as exc:  # noqa: BLE001
                msg = str(exc)
                if ambiguous_policy == "block":
                    raise RuntimeError(f"Lane B failed for item {code}: {msg}") from exc
                errors.append({"item_code": code, "error": "lane_b_failed", "detail": msg})
                parsed.notes.append("lane_b_failed_continue")
                parsed.lane_reason = "lane_b_failed_continue"
        elif use_lane_b and not groq_api_key:
            if ambiguous_policy == "block":
                raise RuntimeError(f"Lane B required but GROQ_API_KEY missing for item {code}")
            parsed.notes.append("lane_b_needed_but_no_api_key")
            parsed.lane_reason = lane_b_reason or "fallback_needed_no_key"
        parsed = enforce_hard_rules(parsed)
        if not instruction_meta.get("has_action_signal", True):
            parsed.notes.append("instruction_without_action_signal")
        per_child[code] = parsed

    parent_rows: Dict[str, dict] = {}
    for code in child_order:
        child = children[code]
        parent = child.parent
        if parent not in parent_rows:
            p = parents.get(parent, ParentItem(parent, f"Parent {parent}", child.unit))
            parent_rows[parent] = {
                "code": parent,
                "description": p.description,
                "unit": p.unit or child.unit,
                "quantity": 0.0,
                "remove": 0.0,
                "dispose": 0.0,
                "reuse": 0.0,
                "clean": 0.0,
                "package": 0.0,
                "transport": 0.0,
                "children": [],
            }
        row = parent_rows[parent]
        row["children"].append(code)
        row["quantity"] += child.quantity

        parsed = per_child.get(code)
        if not parsed:
            continue

        reuse_amt = compute_amount(child.quantity, parsed.reuse, 0.0)
        dispose_amt = compute_amount(child.quantity, parsed.dispose, reuse_amt)
        remove_amt = compute_amount(child.quantity, parsed.remove, reuse_amt)
        clean_amt = compute_amount(child.quantity, parsed.clean, reuse_amt)
        package_amt = compute_amount(child.quantity, parsed.package, reuse_amt)
        transport_amt = compute_amount(child.quantity, parsed.transport, reuse_amt)

        row["remove"] += remove_amt
        row["dispose"] += dispose_amt
        row["reuse"] += reuse_amt
        row["clean"] += clean_amt
        row["package"] += package_amt
        row["transport"] += transport_amt

    wb = Workbook()
    ws = wb.active
    ws.title = "Blad1"
    ws.append(HEADER)
    bold = Font(bold=True)
    for c in range(1, len(HEADER) + 1):
        ws.cell(row=1, column=c).font = bold

    parent_seen: set[str] = set()
    for code in child_order:
        child = children[code]
        parent = child.parent
        if parent not in parent_seen:
            parent_seen.add(parent)
            prow = parent_rows[parent]
            ws.append(
                [
                    float(parent),
                    prow["description"],
                    prow["unit"],
                    round2(prow["quantity"]),
                    blank_if_zero(prow["remove"]),
                    blank_if_zero(prow["dispose"]),
                    blank_if_zero(prow["reuse"]),
                    blank_if_zero(prow["clean"]),
                    blank_if_zero(prow["package"]),
                    blank_if_zero(prow["transport"]),
                ]
            )
            prow_idx = ws.max_row
            for c in range(1, len(HEADER) + 1):
                ws.cell(row=prow_idx, column=c).font = bold
        ws.append(
            [
                float(code),
                child.description,
                child.unit,
                round2(child.quantity),
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )

    validation_issues: List[dict] = []
    for code in child_order:
        if code not in per_child:
            validation_issues.append({"item_code": code, "type": "missing_parse"})
            continue
        p = per_child[code]
        if not (0.0 - EPS <= p.reuse.pct <= 100.0 + EPS):
            validation_issues.append({"item_code": code, "type": "reuse_out_of_range"})
        if not (0.0 - EPS <= p.dispose.pct <= 100.0 + EPS):
            validation_issues.append({"item_code": code, "type": "dispose_out_of_range"})
        if p.reuse.pct + p.dispose.pct > 100.0 + EPS:
            validation_issues.append({"item_code": code, "type": "reuse_dispose_over_100"})

    for parent, row in parent_rows.items():
        qty_children = sum(children[c].quantity for c in row["children"])
        if abs(qty_children - row["quantity"]) > EPS:
            validation_issues.append({"item_code": parent, "type": "parent_quantity_mismatch"})

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)

    report = {
        "summary": {
            "children_in_file1": len(child_order),
            "instructions_found": len(instructions),
            "parsed_children": len(per_child),
            "errors": len(errors),
            "lane_a": sum(1 for p in per_child.values() if p.lane == "A"),
            "lane_b": sum(1 for p in per_child.values() if p.lane.startswith("B")),
            "validation_issues": len(validation_issues),
        },
        "policies": {
            "canonical_verb_dictionary": {
                "reuse": ["reuse", "reused", "reused locally"],
                "dispose": ["dispose", "disposed", "disposed of"],
                "clean": ["clean", "cleaned", "fully cleaned"],
                "package": ["package", "packaged"],
                "transport": ["transport costs", "transport"],
            },
            "rounding_policy": "round_half_up_to_2_decimals_for_output",
            "rounding_policy_detail": "decimal ROUND_HALF_UP at final output cell write",
            "conflict_policy": (
                "lane_a first; if low confidence or inconsistent percentages, lane_b with strict schema"
            ),
            "missing_instruction_policy": (
                "keep processing; mark error in report; child contributes to quantity but no action allocation"
            ),
            "versioning_policy": (
                "persist parser version + model in report; update patterns in source-controlled releases"
            ),
        },
        "children": [asdict(p) for p in per_child.values()],
        "instruction_trace": {
            k: {
                "page": v.get("page"),
                "block_hash": v.get("block_hash"),
                "has_action_signal": v.get("has_action_signal"),
            }
            for k, v in instructions.items()
        },
        "errors": errors,
        "validation_issues": validation_issues,
        "runtime": {
            "groq_model": groq_model,
            "parser_version": "v1.0.0",
            "force_lane_b": force_lane_b,
            "ambiguous_policy": ambiguous_policy,
        },
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")


def main() -> int:
    base_dir = Path(__file__).resolve().parent
    load_dotenv(base_dir / ".env")
    parser = argparse.ArgumentParser(
        description="Generate Part 2 output from File 1 and File 2."
    )
    parser.add_argument(
        "--file1",
        default=str(base_dir / "inputs/[FILE 1]_ subset of registration status.xlsx"),
        help="Input Excel file (File 1)",
    )
    parser.add_argument(
        "--file2",
        default=str(base_dir / "inputs/FILE 2 subset.pdf"),
        help="Input PDF file (File 2)",
    )
    parser.add_argument(
        "--output",
        default=str(base_dir / "outputs/FILE_3_GENERATED.xlsx"),
        help="Output Excel path",
    )
    parser.add_argument(
        "--report",
        default=str(base_dir / "reports/part2_run_report.json"),
        help="Run report path",
    )
    parser.add_argument(
        "--groq-model",
        default=os.getenv("GROQ_MODEL", "openai/gpt-oss-20b"),
        help="Groq model for lane B fallback",
    )
    parser.add_argument(
        "--force-lane-b",
        action="store_true",
        help="Force all items through Groq fallback (for integration testing).",
    )
    parser.add_argument(
        "--ambiguous-policy",
        choices=["continue", "block"],
        default="continue",
        help="Behavior when ambiguous instruction cannot be resolved (default: continue + flag).",
    )
    args = parser.parse_args()

    groq_api_key = os.getenv("GROQ_API_KEY")

    try:
        process(
            file1=Path(args.file1),
            file2_pdf=Path(args.file2),
            output_xlsx=Path(args.output),
            report_path=Path(args.report),
            groq_model=args.groq_model,
            groq_api_key=groq_api_key,
            force_lane_b=args.force_lane_b,
            ambiguous_policy=args.ambiguous_policy,
        )
    except Exception as exc:  # noqa: BLE001
        sys.stderr.write(f"ERROR: {exc}\n")
        return 1

    print(
        textwrap.dedent(
            f"""
            Success.
            Output: {args.output}
            Report: {args.report}
            """
        ).strip()
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
